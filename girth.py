import openpyxl,sys
from openpyxl import *



def stringMSJ( input ) :
	"This will parse the input string"
	parsedString = input.split(",")
	return parsedString

def highestRow( ws ) :
	cellChecker = False
	rowNum = 1
	startCol = 'A'

	while(cellChecker == False) :
		if(ws[startCol + str(rowNum)].value == None) :
			cellChecker = True
			print("Row " + str(rowNum) + " is empty")
		else :
			print("Row " + str(rowNum) + " is not empty")
			rowNum += 1
	return rowNum
				

def excellInputer( parsedString ) :	
	teamNum = None
	matchNum = None
	alliance = None	
	startPosition = None
	placedGear = None
	scoredHG = None
	scoredLG = None
	numGear = None
	numLG = None
	numHG = None
	climbed = None
	matchNotes = None
	startRow = 2
	startCol = 'A'
	tableX = 12
	tableY = 6
	countTeam = False
	countMatch = False
	teamList = 'WPITeamList.xlsx'
	matchList = 'WPIMatchList.xlsx'
	
	try:
		
		output = stringMSJ(parsedString)
		length = len(output)
			
		if(length >= 1) : teamNum = output[0]
		if(length >= 2) : matchNum = output[1]
		if(length >= 3) : alliance = output[2]
		if(length >= 4) : startPosition = output[3]
		if(length >= 5) : placedGear = output[4]
		if(length >= 6) : scoredHG = output[5]
		if(length >= 7) : scoredLG = output[6]
		if(length >= 8) : numGear = output[7]
		if(length >= 9) : numLG = output[8]
		if(length >= 10) : numHG = output[9]
		if(length >= 11) : climbed = output[10]
		if(length >= 12) : matchNotes = output[11]



		wbMatch = load_workbook(matchList)
		wbTeam = load_workbook(teamList)

		templateMatch = wbMatch.active
		templateTeam = wbTeam.active
		
		for sheet in wbMatch :
			if(sheet.title == matchNum) :
				countMatch = True
		
		for sheet in wbTeam :
			if(sheet.title == teamNum):
				countTeam = True
			
			
		if(not countMatch) : 
			wsMatch = wbMatch.copy_worksheet(templateMatch)
			wsMatch.title = (matchNum)
			print(wsMatch.title + " is not in MatchList.")
			print(wbMatch.sheetnames)
		elif(countMatch) :
			wsMatch = wbMatch[matchNum]
			print(wsMatch.title + " is in MatchList.")
			print(wbMatch.sheetnames)
			
		if(not countTeam) : 
			wsTeam = wbTeam.copy_worksheet(templateTeam)
			wsTeam.title = (teamNum)
			print(wsTeam.title + " is not in TeamList.")
			print(wbTeam.sheetnames)
		elif(countTeam) :
			wsMatch = wbTeam[teamNum]
			print(wsTeam.title + " is in TeamList.")
			print(wbTeam.sheetnames)
			
		print("Match List")
		rowNumMatch = highestRow( wsMatch )
		
		print("Team List")
		rowNumTeam = highestRow( wsTeam )
		
		wsMatch[startCol + str(rowNumMatch)].value = teamNum
		wsTeam[startCol + str(rowNumTeam)].value = matchNum
		
		wsMatch['B' + str(rowNumMatch)].value = alliance
		wsTeam['B' + str(rowNumTeam)].value = alliance
		wsMatch['C' + str(rowNumMatch)].value = startPosition
		wsTeam['C' + str(rowNumTeam)].value = startPosition
		wsMatch['D' + str(rowNumMatch)].value = placedGear
		wsTeam['D' + str(rowNumTeam)].value = placedGear
		wsMatch['E' + str(rowNumMatch)].value = scoredHG
		wsTeam['E' + str(rowNumTeam)].value = scoredHG
		wsMatch['F' + str(rowNumMatch)].value = scoredLG
		wsTeam['F' + str(rowNumTeam)].value = scoredLG
		wsMatch['G' + str(rowNumMatch)].value = numGear
		wsTeam['G' + str(rowNumTeam)].value = numGear
		wsMatch['H' + str(rowNumMatch)].value = numLG
		wsTeam['H' + str(rowNumTeam)].value = numLG
		wsMatch['I' + str(rowNumMatch)].value = numHG
		wsTeam['I' + str(rowNumTeam)].value = numHG
		wsMatch['J' + str(rowNumMatch)].value = climbed
		wsTeam['J' + str(rowNumTeam)].value = climbed
		wsMatch['K' + str(rowNumMatch)].value = matchNotes
		wsTeam['K' + str(rowNumTeam)].value = matchNotes


		wbMatch.save(matchList)
		wbTeam.save(teamList)
		
		return True
	except Exception as e :
		print(e)
		return e
	















	











