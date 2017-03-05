import openpyxl,xlsxwriter,sys
from openpyxl import *



def stringMSJ( input ) :
	"This will parse the input string"
	parsedString = input.split(",")
	return parsedString

def excellInputer( parsedString ) :	
	teamNum = None
	matchNum = None
	alliance = None
	startPosition = None
	placedGear = None
	scoredHG = None
	scoredLG = None
	numGear = None
	numLG = None
	numHG = None
	climbed = None
	matchNotes = None
	startRow = 2
	startCol = 'A'
	tableX = 12
	tableY = 6
	count = False
	swag = "35"
	destfile_name = 'WPITable.xlsx'
	
	try:
		
		output = stringMSJ(parsedString)
		length = len(output)
			
		if(length >= 1) : teamNum = output[0]
		if(length >= 2) : matchNum = output[1]
		if(length >= 3) : alliance = output[2]
		if(length >= 4) : startPosition = output[3]
		if(length >= 5) : placedGear = output[4]
		if(length >= 6) : scoredHG = output[5]
		if(length >= 7) : scoredLG = output[6]
		if(length >= 8) : numGear = output[7]
		if(length >= 9) : numLG = output[8]
		if(length >= 10) : numHG = output[9]
		if(length >= 11) : climbed = output[10]
		if(length >= 12) : matchNotes = output[11]



		wb = load_workbook(destfile_name)

		template = wb.active

		for sheet in wb:
			if(sheet.title == matchNum) :
				count = True
			
			
		if(count == False) : 
			ws = wb.copy_worksheet(template)
			ws.title = (matchNum)
			print(ws.title + " is not here.")
			print(wb.sheetnames)
		elif(count == True) :
			ws = wb[matchNum]
			print(ws.title + " is here.")
			print(wb.sheetnames)
			
		cell = 'A2'
		if(ws[cell].value == None) : print("oh no")

		cellChecker = False
		rowNum = 1

		while(cellChecker == False) :
			if(ws[startCol + str(rowNum)].value == None) :
				cellChecker = True
				print("Row " + str(rowNum) + " is empty")
				ws[startCol + str(rowNum)] = teamNum
			else :
				print("Row " + str(rowNum) + " is not empty")
				rowNum += 1
				
		ws['B' + str(rowNum)].value = alliance
		ws['C' + str(rowNum)].value = startPosition
		ws['D' + str(rowNum)].value = placedGear
		ws['E' + str(rowNum)].value = scoredHG
		ws['F' + str(rowNum)].value = scoredLG
		ws['G' + str(rowNum)].value = numGear
		ws['H' + str(rowNum)].value = numLG
		ws['I' + str(rowNum)].value = numHG
		ws['J' + str(rowNum)].value = climbed
		ws['K' + str(rowNum)].value = matchNotes


		wb.save(destfile_name)
		
		return True
	except Exception as e :
		print(e)
		return e
	















	











